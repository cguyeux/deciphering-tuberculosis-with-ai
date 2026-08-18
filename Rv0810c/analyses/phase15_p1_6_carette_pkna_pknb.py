#!/usr/bin/env python3
"""
P1.6 — la phosphorylation de T24 (Rv0810c) baisse-t-elle sous inhibition de
PknA/PknB (Carette et al. 2018, mBio, PMID 29511081) ?
Verification sur piece de la Table S3 (mbo001183756st3.xlsx, Europe PMC
supplementaryFiles), pas seulement sur le resume d'un agent de recherche.

Limite de l'experience a garder en tete pour l'interpretation : l'inhibiteur
est un pyrimidine substitue actif sur PknA (Ki=0,018 uM) ET PknB (Ki=0,004 uM)
simultanement -> ne peut PAS distinguer laquelle des deux kinases phosphoryle
T24, seulement etablir une dependance jointe PknA/PknB.
"""
import json
from pathlib import Path

import openpyxl

XLSX = Path(__file__).resolve().parent.parent / "experiments" / "2026-08-10_P1_6_carette2018" / "mbo001183756st3.xlsx"
OUT_JSON = Path(__file__).resolve().parent.parent / "résultats" / "p1_6_carette_pkna_pknb.json"

TARGET_PEPTIDE_SUBSTR = "YSSPQT"  # peptide YSSPQTDFQR, Rv0810c 19-28, phospho sur T24


def find_rv0810c_rows(sheet):
    """Convention constante sur les 6 onglets S3A-S3F : ligne 0 = titre, ligne 1 = vide,
    ligne 2 = en-tetes de colonnes, ligne 3+ = donnees (verifie manuellement sur S3C/S3F)."""
    all_rows = list(sheet.iter_rows(values_only=True))
    if len(all_rows) < 4:
        return []
    header = [str(c) if c is not None else "" for c in all_rows[2]]
    rows = []
    for row in all_rows[3:]:
        row_str = [str(c) if c is not None else "" for c in row]
        if any("Rv0810c" in c for c in row_str) or any(TARGET_PEPTIDE_SUBSTR in c for c in row_str):
            rows.append(dict(zip(header, row)))
    return rows


def main():
    wb = openpyxl.load_workbook(str(XLSX), read_only=True, data_only=True)
    out = {}
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = find_rv0810c_rows(ws)
        out[sheet_name] = rows
        print(f"=== {sheet_name}: {len(rows)} ligne(s) Rv0810c/T24 ===")
        for r in rows:
            print({k: v for k, v in r.items() if v not in (None, "")})

    OUT_JSON.write_text(json.dumps(out, indent=2, ensure_ascii=False, default=str))


if __name__ == "__main__":
    main()
