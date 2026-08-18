#!/usr/bin/env python3
"""
P1.3 — INSTRUIRE le contre-argument CRISPRi de Choudhery et al. 2024.

Contexte. Le projet a delibérement abandonne l'essentialite TnSeq (5 sites TA
dont 4 permissifs) pour s'appuyer sur la vulnerabilite CRISPRi de Bosch 2021
(VI = -11,98 [-15,41 ; -8,60]). Choudhery, DeJesus, Srinivasan, Rock,
Schnappinger & Ioerger 2024 (PLoS Comput Biol, PMID 38768228) — le groupe meme
d'ou vient la bibliotheque RLC12 — utilise Rv0810c comme EXEMPLE PUBLIE de gene
CRISPRi bruite. Le contre-argument vise donc exactement le pilier restant.

Fait decouvert le 2026-08-10 en relisant la fiche atlas, et VERIFIE ici sur la
table brute : Bosch et al. 2021 eux-memes classent Rv0810c `certain = False`,
c'est-a-dire que le gene ne remplit PAS leurs criteres d'estimation fiable de
la vulnerabilite. Ce n'est donc pas un desaccord entre deux equipes : c'est le
meme groupe qui, deux fois et par deux voies, signale ce gene comme non fiable.

La lecture charitable a instruire (elle est dans la piste, elle n'est pas
supposee) : un gene TRES vulnerable est deplete avant meme le debut de
l'experience, donc peu abondant et bruite ; le bruit serait la CONSEQUENCE de
la vulnerabilite, pas la preuve de son artefact. Cette hypothese fait une
prediction falsifiable : les genes a VI tres negatif devraient etre
majoritairement `certain=False`. Si au contraire les genes `certain=True`
portent les VI les plus extremes, l'hypothese tombe et le flag de Rv0810c
parle de la QUALITE DE L'AJUSTEMENT, pas de sa vulnerabilite.

Trois questions de la piste, plus celle-ci :
  (1) la faible abondance de base est-elle presente aussi SANS drogue ?
      -> proxy local : taux de sgRNA "flatliner" et parametres du two-line fit
         de Bosch, mesures en l'absence de tout antibiotique.
  (2) le bruit est-il propre a ce gene, ou generique aux genes tres vulnerables ?
      -> comparaison a un panel apparie sur le VI et sur n_guides.
  (3) reconcilier 23 guides comptes (RLC12) / 22 rapportes (Choudhery).

Sortie : résultats/p1_3_crispri_counterargument.json
"""

from __future__ import annotations

import json
from pathlib import Path

import numpy as np
import openpyxl

PROJ = Path(__file__).resolve().parent.parent
BOSCH = PROJ.parent / "annotation_mtbc" / "data" / "bosch2021"
MMC3 = BOSCH / "mmc3.xlsx"
MMC2 = BOSCH / "mmc2.xlsx"
OUTDIR = PROJ / "résultats"
OUTDIR.mkdir(exist_ok=True)

TARGET = "RVBD0810c"


def norm(tag: str) -> str:
    return str(tag).strip()


def load_genes() -> tuple[list[str], dict[str, dict]]:
    wb = openpyxl.load_workbook(MMC3, read_only=True)
    ws = wb["(1) Mtb H37Rv"]
    it = ws.iter_rows(values_only=True)
    hdr = [str(h) for h in next(it)]
    out: dict[str, dict] = {}
    for r in it:
        if r[0] is None:
            continue
        d = dict(zip(hdr, r))
        out[norm(d["locus_tag"])] = d
    return hdr, out


def load_guides() -> dict[str, list[dict]]:
    wb = openpyxl.load_workbook(MMC2, read_only=True)
    ws = wb["(4) Mtb H37Rv two-line fits"]
    it = ws.iter_rows(values_only=True)
    hdr = [str(h) for h in next(it)]
    by_gene: dict[str, list[dict]] = {}
    for r in it:
        if r[0] is None:
            continue
        d = dict(zip(hdr, r))
        # "Gene" a la forme "RVBD0810c:RVBD0810c" ou "16S:16S"
        g = str(d["Gene"]).split(":")[0]
        by_gene.setdefault(g, []).append(d)
    return by_gene


def fnum(v) -> float:
    try:
        return float(v)
    except (TypeError, ValueError):
        return float("nan")


def fbool(v) -> bool:
    return str(v).strip().lower() in ("true", "1", "vrai")


def describe(a: np.ndarray) -> dict:
    a = a[~np.isnan(a)]
    if len(a) == 0:
        return {"n": 0}
    return {
        "n": int(len(a)),
        "mediane": round(float(np.median(a)), 3),
        "moyenne": round(float(np.mean(a)), 3),
        "q05": round(float(np.percentile(a, 5)), 3),
        "q95": round(float(np.percentile(a, 95)), 3),
    }


def main() -> None:
    hdr, genes = load_genes()
    tgt = genes[TARGET]
    rep: dict = {
        "piste": "P1.3",
        "source": str(MMC3),
        "cible_Rv0810c_ligne_brute": {
            k: (str(v) if not isinstance(v, (int, float)) or v is None else v)
            for k, v in tgt.items()
        },
    }

    tags = list(genes)
    vi = np.array([fnum(genes[t]["Vulnerability Index"]) for t in tags])
    cert = np.array([fbool(genes[t]["certain"]) for t in tags])
    ng = np.array([fnum(genes[t]["n_guides"]) for t in tags])
    sspan = np.array([fnum(genes[t]["str_span"]) for t in tags])
    mspan = np.array([fnum(genes[t]["M_span"]) for t in tags])
    bspan = np.array([fnum(genes[t]["beta_max_span"]) for t in tags])
    hwidth = np.array([fnum(genes[t]["u_H"]) - fnum(genes[t]["l_H"]) for t in tags])
    viwidth = np.array(
        [fnum(genes[t]["VI Upper Bound"]) - fnum(genes[t]["VI Lower Bound"]) for t in tags]
    )
    crisp_ess = np.array([str(genes[t]["crispr_ess"]) == "Essential" for t in tags])
    ti = tags.index(TARGET)

    # ------------------------------------------------------------------ #
    # 1. Le flag `certain` : que vaut-il a l'echelle du jeu ?
    # ------------------------------------------------------------------ #
    rep["1_flag_certain"] = {
        "definition_Bosch": (
            "certain = le gene remplit les criteres d'une estimation FIABLE de "
            "la vulnerabilite (legende de mmc3.xlsx)."
        ),
        "Rv0810c_certain": bool(cert[ti]),
        "n_genes": len(tags),
        "n_certain_True": int(cert.sum()),
        "n_certain_False": int((~cert).sum()),
        "VI_si_certain_True": describe(vi[cert]),
        "VI_si_certain_False": describe(vi[~cert]),
    }

    # ------------------------------------------------------------------ #
    # 2. LE TEST QUI TRANCHE la lecture charitable.
    #    Hypothese : "tres vulnerable -> deplete -> bruite -> certain=False".
    #    Prediction : les VI les plus negatifs seraient majoritairement
    #    certain=False. Verification directe.
    # ------------------------------------------------------------------ #
    bins = [(-1e9, -15), (-15, -10), (-10, -5), (-5, 0), (0, 1e9)]
    tab = []
    for lo, hi in bins:
        m = (vi > lo) & (vi <= hi) & ~np.isnan(vi)
        n = int(m.sum())
        tab.append(
            {
                "classe_VI": f"({lo if lo > -1e8 else '-inf'} ; {hi if hi < 1e8 else '+inf'}]",
                "n": n,
                "n_certain_True": int(cert[m].sum()),
                "pct_certain_True": round(100.0 * cert[m].mean(), 1) if n else None,
            }
        )
    rep["2_test_de_la_lecture_charitable"] = {
        "hypothese": (
            "Si le flag `certain=False` etait la consequence d'une vulnerabilite "
            "extreme (deplétion trop rapide), alors les genes a VI tres negatif "
            "seraient majoritairement non-certains."
        ),
        "table_certain_par_classe_de_VI": tab,
        "VI_Rv0810c": round(float(vi[ti]), 3),
        "n_genes_VI_inferieur_ou_egal_a_Rv0810c": int(np.nansum(vi <= vi[ti])),
        "parmi_eux_n_certain_True": int(cert[(vi <= vi[ti]) & ~np.isnan(vi)].sum()),
    }

    # ------------------------------------------------------------------ #
    # 3. Ce qui distingue REELLEMENT Rv0810c : la largeur des intervalles.
    # ------------------------------------------------------------------ #
    def rank_pct(val: float, pop: np.ndarray, among: np.ndarray) -> dict:
        a = pop[among & ~np.isnan(pop)]
        return {
            "valeur_Rv0810c": round(float(val), 3),
            "n_comparaison": int(len(a)),
            "mediane": round(float(np.median(a)), 3),
            "percentile_de_Rv0810c": round(100.0 * float(np.mean(a <= val)), 1),
        }

    ess_mask = crisp_ess & ~np.isnan(vi)
    rep["3_qualite_de_l_ajustement"] = {
        "reference": "parmi les genes appeles Essential par CRISPRi",
        "str_span": rank_pct(sspan[ti], sspan, ess_mask),
        "M_span": rank_pct(mspan[ti], mspan, ess_mask),
        "beta_max_span": rank_pct(bspan[ti], bspan, ess_mask),
        "largeur_IC_H": rank_pct(hwidth[ti], hwidth, ess_mask),
        "largeur_IC_VI": rank_pct(viwidth[ti], viwidth, ess_mask),
        "n_guides": rank_pct(ng[ti], ng, ess_mask),
    }

    # ------------------------------------------------------------------ #
    # 4. Niveau sgRNA : les "flatliners" (mesure SANS aucune drogue)
    # ------------------------------------------------------------------ #
    guides = load_guides()
    g_t = guides.get(TARGET, [])
    flat_frac: dict[str, float] = {}
    nguides_obs: dict[str, int] = {}
    for g, lst in guides.items():
        nguides_obs[g] = len(lst)
        flat_frac[g] = float(np.mean([fbool(d["Flat"]) for d in lst])) if lst else np.nan

    ff = np.array([flat_frac.get(t, np.nan) for t in tags])
    # panel apparie : genes essentiels a VI comparable (+/- 3) et n_guides +/- 8
    matched = (
        crisp_ess
        & (np.abs(vi - vi[ti]) <= 3)
        & (np.abs(ng - ng[ti]) <= 8)
        & ~np.isnan(ff)
    )
    matched_certain = matched & cert
    rep["4_niveau_sgRNA_flatliners"] = {
        "definition_Flat": (
            "sgRNA classe 'flatliner' par la regression par morceaux de Bosch : "
            "aucune deplétion detectable. Mesure faite SANS antibiotique, ce qui "
            "repond directement a la question (1) de la piste."
        ),
        "Rv0810c": {
            "n_sgRNA_dans_le_fit": len(g_t),
            "n_guides_declare_mmc3": int(ng[ti]),
            "n_flatliners": int(sum(fbool(d["Flat"]) for d in g_t)),
            "frac_flatliners": round(float(ff[ti]), 3) if not np.isnan(ff[ti]) else None,
            "Alpha_L_median": round(
                float(np.nanmedian([fnum(d["Alpha_L"]) for d in g_t])), 3
            )
            if g_t
            else None,
            "Beta_E_median": round(
                float(np.nanmedian([fnum(d["Beta_E"]) for d in g_t])), 3
            )
            if g_t
            else None,
        },
        "fond_tous_genes": describe(ff),
        "fond_essentiels_CRISPRi": describe(ff[crisp_ess]),
        "panel_apparie_VI_et_n_guides": {
            "critere": f"Essential, |VI - {vi[ti]:.2f}| <= 3, |n_guides - {int(ng[ti])}| <= 8",
            "n": int(matched.sum()),
            "frac_flatliners": describe(ff[matched]),
            "percentile_de_Rv0810c": round(
                100.0 * float(np.mean(ff[matched] <= ff[ti])), 1
            )
            if matched.sum()
            else None,
            "dont_certain_True": int(matched_certain.sum()),
            "frac_flatliners_si_certain_True": describe(ff[matched_certain]),
        },
    }

    # ------------------------------------------------------------------ #
    # 5. Reconciliation 23 / 22 guides
    # ------------------------------------------------------------------ #
    rep["5_reconciliation_comptage_guides"] = {
        "RLC12_bibliotheque_P0.3": 23,
        "mmc3_n_guides": int(ng[ti]),
        "mmc2_sgRNA_effectivement_ajustes": len(g_t),
        "Choudhery2024_rapporte": 22,
        "lecture": (
            "Le comptage bibliotheque (23) et le comptage Bosch (mmc3) doivent "
            "coincider ; l'ecart avec Choudhery se joue au filtrage de son propre "
            "jeu (un sgRNA elimine pour comptages insuffisants dans l'experience "
            "RIF D10), pas sur la composition de la bibliotheque."
        ),
    }

    out = OUTDIR / "p1_3_crispri_counterargument.json"
    out.write_text(json.dumps(rep, indent=2, ensure_ascii=False, default=str))
    print(json.dumps(rep, indent=2, ensure_ascii=False, default=str))
    print(f"\n[ecrit] {out}")


if __name__ == "__main__":
    main()
